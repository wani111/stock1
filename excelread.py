import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

 
#data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("C:\\Users\\kingw\\OneDrive\\stock\\roe_chart_2.xlsx", data_only=True)
#시트 이름으로 불러오기
load_ws = load_wb['Total']
 
def get_bs_obj(company_code):
    url = "https://finance.naver.com/item/main.nhn?code=" + company_code
    result = requests.get(url)
    bs_obj = BeautifulSoup(result.content, "html.parser")
    return bs_obj

# company_code를 입력받아 now_price를 출력
def get_price(company_code):
    #print("get_price " + company_code)
    bs_obj = get_bs_obj(company_code)
    no_today = bs_obj.find("p", {"class": "no_today"})
    blind = no_today.find("span", {"class": "blind"})
    now_price = blind.text
    return now_price

#excel_stock_code = load_ws.cell(3,9).value
#print(excel_stock_code[1:])
#now_price = get_price(excel_stock_code[1:])
#print(now_price)
#load_ws.cell(1,1,"test")
#load_wb.save("C:\\Users\\kingw\\OneDrive\\stock\\roe_chart.xlsx")
# -------------------------------------------------------#
j = 0
for i in load_ws['I']:
	#print(i)
	j = j + 1
	code = i.value
	if code == None:
		continue
	if code == "종목코드":
		continue
	#print(code)
	now_price = get_price(code[1:])
	print(j, code, now_price)
	load_ws.cell(j,2,now_price)
	load_wb.save("C:\\Users\\kingw\\OneDrive\\stock\\roe_chart.xlsx")	


#load_ws.cell(1,1,"test")
#load_wb.save("C:\\Users\\kingw\\OneDrive\\stock\\roe_chart.xlsx")